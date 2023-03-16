from openpyxl import load_workbook
import time
import datetime

print("STARTING")
count = 1

#open and load file, select OF sheet, search header portion
wb = load_workbook(filename='Acme Furniture 2022 Feb 1- 14.xlsx', data_only=True)
sbs = load_workbook(filename= 'sbs.xlsx', data_only=True)
ws = wb.active
sbs_ws = sbs["1. CBP_40HC"]
sheet_headers = ws["A4:J9"]
fcl_table = ws["A10:J27"]
of_table = ws["L10:AB26"]
dray_table = ws["N10:R26"]
arb_table = ws["X10:AC27"]

sbs_ws = sbs["1. CBP_40HC"]
sbs_header = sbs_ws["A1"]
renew_date = sbs_ws["A2"]
rep_info = sbs_ws["B2:W2"]
carrier_header = sbs_ws["A4"]
carrier_info = sbs_ws["B4:W4"]
effective_date = sbs_ws["A6:W6"]
expiration_date = sbs_ws["A7:W7"]
aws_destinations = sbs_ws["A9:A33"]
aws_code = sbs_ws["B9:B33"]
ripi_destinations = sbs_ws["A35:A58"]
ripi_code = sbs_ws["B35:B58"]
ipi_destinations = sbs_ws["A60:A117"]
ipi_code = sbs_ws["B60:B117"]
comm_bullet = sbs_ws["A8:W8"]
aws_cost = sbs_ws["C4:W33"]
ripi_cost = sbs_ws["C4:W58"]
ipi_cost = sbs_ws["C4:W117"]


#list of header dictionaries

header_list = ["Attn:", "Company:", "Company Email:", "From:", "Date:", "Email:", "EFFECTIVE GATE IN:", "to", "COMMODITY:"]

header_dict = {"Attn": None, "Company": None, "Company_Email": None, "From": None, "Date": None, "Email": None, "Effective_Gate_In": None, "Expires": None, "COMMODITY": None}
dray_dict = {"Chassis": [], "I/H": [], "I/H Inclusive notes": [], "Live unload / Drop Pick": [], "Drayage quote# / expire date": []}
of_dict = {"CARRIERS": [], "CY or RAMP": [], "COST 20'": [], "COST 40'": [], "COST 40HQ": [], "COST 45'": [], "PROFIT": [], "Other Per Cntr": []}
arb_dict = {"ARB COST 20'": [], "ARB COST 40'": [], "ARB COST 40HQ": [], "ARB COST 45'": [], "REMARKS": []}
sell_dict = {"ORIGIN": [], "DESTINATION": [], "MODE": [], "TIER": [], "CHASSIS": [], "20'": [], "40'": [], "HQ'": [], "45'": [], "REMARKS": []}

sbs_header_dict = {"Sheet": None}
rep_dict = {"Rep":[]}
carrier_dict = {"Carrier": []}
renew_date_dict = {"Renewal Date": None}
effect_date_dict = {"Effective Date": []}
expire_date_dict = {"Expiration Date": []}
comm_bullet_dict = {"Comm Bullet": []}
aws_dest_dict = {"AWS Destinations": []}
aws_code_dict = {"AWS Code": []}
ripi_dest_dict = {"Ripi Destinations": []}
ripi_code_dict = {"Ripi Code": []}
ipi_dest_dict = {"IPI Destinations": []}
ipi_code_dict = {"IPI Code": []}
aws_cost_dict = {"CMA": [], "CMA-2": [], "COSCO":[], "EMC":[], "Hapag":[], "HMM":[], "HMM-2":[], "MSC":[], "OOCL":[], "ONE":[], "SM Line":[], "YML":[], "ZIM":[], "WHL":[], "WHL-2":[], "Matson":[], "CMA EXX":[], "CULine":[], "CULine-2":[], "CULine-3":[], "Transfar": [], "SeaLead":[]}
ripi_cost_dict = {"CMA": [], "CMA-2": [], "COSCO":[], "EMC":[], "Hapag":[], "HMM":[], "HMM-2":[], "MSC":[], "OOCL":[], "ONE":[], "SM Line":[], "YML":[], "ZIM":[], "WHL":[], "WHL-2":[], "Matson":[], "CMA EXX":[], "CULine":[], "CULine-2":[], "CULine-3":[], "Transfar": [], "SeaLead":[]}
ipi_cost_dict = {"CMA": [], "CMA-2": [], "COSCO":[], "EMC":[], "Hapag":[], "HMM":[], "HMM-2":[], "MSC":[], "OOCL":[], "ONE":[], "SM Line":[], "YML":[], "ZIM":[], "WHL":[], "WHL-2":[], "Matson":[], "CMA EXX":[], "CULine":[], "CULine-2":[], "CULine-3":[], "Transfar": [], "SeaLead":[]}
def function_one(cell, dict, count):
    str_cell = str(cell)
    #print(str_cell)
    #Convert col letter to int equivalent
    col_pos = ord(str_cell[-3])

    #get row position
    row_pos = str_cell[-2]

    #get next column position
    next_col_pos = chr(col_pos + count)

    #get next cell position
    next_cell = sbs_ws[(next_col_pos + row_pos)]

    #get next cell value
    cell_value = next_cell.value
    #print(cell_value)
    dict[key].append(cell_value)

def function_two(cell, dict, count):
    str_cell = str(cell)

    col_pos = ord(str_cell[-3])

    row_pos = str_cell[-2]

    next_col_pos = chr(col_pos + count)

    next_cell = sbs_ws[(next_col_pos + row_pos)]

    cell_value = next_cell.value
        #print(cell_value)
    if cell_value != None:
        str_cell_value = str(cell_value)
        dict[key].append(str_cell_value[:10])

def function_three(cell, dict, count):
    str_cell = str(cell)
          
    col_pos = ord(str_cell[-3])

    row_pos = str_cell[-2]

    next_col_pos = chr(col_pos + count)

    next_cell = sbs_ws[(next_col_pos + row_pos)]

    cell_value = next_cell.value
       
    if cell_value != None:
        str_cell_value = str(cell_value)
        dict[key].append(str_cell_value)

def function_four(cell, dict, count):
    str_cell = cell
                            
    col_pos = str_cell[-3]

    row_pos = str_cell[-2]

    next_row_pos = str(int(row_pos) + count)

    next_cell = sbs_ws[(col_pos + next_row_pos)]

    cell_value = next_cell.value
        
    dict[key].append(cell_value)

def function_five(cell, dict, count):
    str_cell = cell
      
    col_pos = str_cell[-4]

    row_pos = str_cell[-3:-1]

    next_row_pos = str(int(row_pos) + count)

    next_cell = sbs_ws[(col_pos + next_row_pos)]

    cell_value = next_cell.value
        
    dict[key].append(cell_value)

def function_six(cell, count):
    head_cell = str(cell)
                
    head_col_pos = head_cell[-3]

    head_row_pos = head_cell[-2]

    head_next_row_pos = str(int(head_row_pos) + count)

    head_next_cell = sbs_ws[(head_col_pos + head_next_row_pos)]

    return head_next_cell


for row in sheet_headers:
    for cell in row:
        if cell.value != None and cell.value in header_list:
            #print(cell.value)
            str_cell = str(cell)

                        #Convert col letter to int equivalent
            col_pos = ord(str_cell[-3])

                        #get row position
            row_pos = str_cell[-2]

                        #get next column position
            next_col_pos = chr(col_pos + 1)

                        #get next cell position
            next_cell = ws[(next_col_pos + row_pos)]

                        #returns <CELL '020123'.B5> as string. I think i was going to use this to parse. Not sure anymore.
                        #str_next_cell = str(next_cell)

                        #Store the value of next_cell, which should be value of the header
            cell_value = next_cell.value

            #check whether cell.value matches dictionary key
            for key in header_dict:
                cell_append = cell.value[:-1]
                #case switch statement that checks if cell.value matches "Company_Email" or "Effective_Gate_In" or "Expires"
                if cell.value == "Company Email:":
                    header_dict.update({"Company_Email": next_cell.value})
                elif cell.value == "EFFECTIVE GATE IN:":
                    cell_date_time = str(next_cell.value)
                    header_dict.update({"Effective_Gate_In": cell_date_time})
                elif cell.value == "to":
                    cell_date_time = str(next_cell.value)
                    header_dict.update({"Expires": cell_date_time})
                elif cell_append == key:
                    #print(key)
                    header_dict.update({key: next_cell.value})
                elif cell.value == "Date:":
                    cell_date_time = str(next_cell.value)
                    header_dict.update({"Date": cell_date_time})
            #print(cell_value)
#row stuff

row_amount = 0
row_count = 0
dray_row_amount = 0
arb_row_amount = 0        
for row in fcl_table:
    row_amount += 1
for row in dray_table:
    dray_row_amount += 1
for row in arb_table:
    arb_row_amount += 1

#sell_rate section
for row in fcl_table:
    for cell in row:
        for key in sell_dict:
            for number in range(row_amount - 2):
                if cell.value != None and cell.value == key:
                    #checks for chassis header
                    if key == "CHASSIS":
                        #print("found it")
                        str_cell = str(cell)
                        #print(str_cell)
                        col_pos = str_cell[-4]
                        #print(col_pos)
                        row_pos = str_cell[-3:-1]
                        #print(row_pos)
                        next_row_pos = str(int(row_pos) + count)
                        #print(next_row_pos)
                        next_cell = ws[(col_pos + next_row_pos)]
                        #print(next_cell)
                        cell_value = next_cell.value
                        #print(cell_value)

                        #section below gets the number of chassis days
                        chassis_str_cell = str(cell)
                        #print(chassis_str_cell)
                                            #Convert col letter to int equivalent
                        chassis_col_pos = ord(chassis_str_cell[-4])
                        #print(chassis_col_pos)
                                            #get row position
                        chassis_row_pos = chassis_str_cell[-3:-1]
                        #print(chassis_row_pos)

                        chassis_next_row_pos = str(int(chassis_row_pos) + count)
                        #print(chassis_next_row_pos)
                                            #get next column position
                        chassis_next_col_pos = chr(chassis_col_pos + 1)
                        #print(chassis_next_col_pos)
                                            #get next cell position
                        chassis_next_cell = ws[(chassis_next_col_pos + chassis_next_row_pos)]
                        #print(chassis_next_cell)

                        chassis_cell_value = chassis_next_cell.value[:6]
                        #print(chassis_cell_value)
                        chassis = cell_value +' '+ chassis_cell_value

                        sell_dict[key].append(chassis)
                        count += 1
                    
                    else:
                        #sell_dict[key].append(cell_value)
                        str_cell = str(cell)
                        #print(str_cell)
                        col_pos = str_cell[-4]
                        #print(col_pos)
                        row_pos = str_cell[-3:-1]
                        #print(row_pos)
                        next_row_pos = str(int(row_pos) + count)
                        #print(next_row_pos)
                        next_cell = ws[(col_pos + next_row_pos)]
                        #print(next_cell)
                        cell_value = next_cell.value
                        #print(cell_value)
                        sell_dict[key].append(cell_value)
                        count += 1
                        #print("done")
                
            count = 1

#this section is for cost
for row in of_table:
    for cell in row:
        for key in of_dict:
            for number in range(row_amount - 2):
                if cell.value != None and cell.value == key:
                    #print("done")
                    str_cell = str(cell)
                    #search for "." in str_cell
                    if "." in str_cell:
                        check_str_cell = len(str_cell[str_cell.index(".") + 1:])
                        if check_str_cell > 4:

                            #print(str_cell)
                            col_pos = str_cell[-5:-3]
                            #print(col_pos)
                            row_pos = str_cell[-3:-1]
                            #print(row_pos)
                            next_row_pos = str(int(row_pos) + count)
                            #print(next_row_pos)
                            next_cell = ws[(col_pos + next_row_pos)]
                            #print(next_cell)
                            cell_value = next_cell.value
                            #print(cell_value)
                            of_dict[key].append(cell_value)
                            count += 1
                        else:
                            #print(str_cell)
                            col_pos = str_cell[-4]
                            #print(col_pos)
                            row_pos = str_cell[-3:-1]
                            #print(row_pos)
                            next_row_pos = str(int(row_pos) + count)
                            #print(next_row_pos)
                            next_cell = ws[(col_pos + next_row_pos)]
                            #print(next_cell)
                            cell_value = next_cell.value
                            #print(cell_value)
                            of_dict[key].append(cell_value)
                            count += 1
            count = 1

#this section is for dray
for row in dray_table:
    for cell in row:
        for key in dray_dict:
            for number in range(dray_row_amount - 2):
                if cell.value != None and cell.value == key:
                    str_cell = str(cell)
                    #print(str_cell)
                    col_pos = str_cell[-4]
                    #print(col_pos)
                    row_pos = str_cell[-3:-1]
                    #print(row_pos)
                    next_row_pos = str(int(row_pos) + count)
                    #print(next_row_pos)
                    next_cell = ws[(col_pos + next_row_pos)]
                    #print(next_cell)
                    cell_value = next_cell.value
                    #print(cell_value)
                    dray_dict[key].append(cell_value)
                    count += 1
            count = 1

#this section is for arb
for row in arb_table:
    for cell in row:
        for key in arb_dict:
            for number in range(arb_row_amount - 2):
                if cell.value != None and cell.value == key:
                    #print("done")
                    str_cell = str(cell)
                    #search for "." in str_cell
                    if "." in str_cell:
                        check_str_cell = len(str_cell[str_cell.index(".") + 1:])
                        if check_str_cell > 4:

                            #print(str_cell)
                            col_pos = str_cell[-5:-3]
                            #print(col_pos)
                            row_pos = str_cell[-3:-1]
                            #print(row_pos)
                            next_row_pos = str(int(row_pos) + count)
                            #print(next_row_pos)
                            next_cell = ws[(col_pos + next_row_pos)]
                            #print(next_cell)
                            cell_value = next_cell.value
                            #print(cell_value)
                            arb_dict[key].append(cell_value)
                            count += 1
                        else:
                            #print(str_cell)
                            col_pos = str_cell[-4]
                            #print(col_pos)
                            row_pos = str_cell[-3:-1]
                            #print(row_pos)
                            next_row_pos = str(int(row_pos) + count)
                            #print(next_row_pos)
                            next_cell = ws[(col_pos + next_row_pos)]
                            #print(next_cell)
                            cell_value = next_cell.value
                            #print(cell_value)
                            arb_dict[key].append(cell_value)
                            count += 1
            count = 1

if sbs_header.value != None:
    header = sbs_header.value
    sbs_header_dict.update({"Sheet": header})

#this gets the renewal date
if renew_date.value != None:

    renew = renew_date.value[9:]
    renew_date_dict.update({"Renewal Date": renew})

#this gets the rep info
for row in rep_info:
    for cell in row:
        for key in rep_dict:
            if cell.value != None:
                function_one(cell, rep_dict, count)
                count += 1
                    
        count = 1

#this gets the carrier info, could probably rewrite this to be more efficient
carrier_header_value = carrier_header.value
if carrier_header_value != None and carrier_header_value == "Carrier":
    for row in carrier_info:
        for cell in row:
            for key in carrier_dict:
                if cell.value == None:
                    function_one(cell, carrier_dict, count)
                    count += 1
                else:
                    function_one(cell, carrier_dict, count)
                    count += 1
                        
            count = 1

#this gets the effective date
for row in effective_date:
    for cell in row:
        for key in effect_date_dict:
        
            function_two(cell, effect_date_dict, count)
            count += 1
                    
        count = 1

#this gets the expiration date
for row in expiration_date:
    for cell in row:
        for key in expire_date_dict:
        
            function_two(cell, expire_date_dict, count)
            count += 1
                    
        count = 1

#this gets the comm bullet
for row in comm_bullet:
    for cell in row:
        for key in comm_bullet_dict:
        
            function_three(cell, comm_bullet_dict, count)
            count += 1
                    
        count = 1


aws_dest_row_count = 0
aws_dest_count = 0
for row in aws_destinations:
    aws_dest_row_count += 1

#this gets the AWS destinations
for row in aws_destinations:
    for cell in row:
        for key in aws_dest_dict:
            if cell.value != None and cell.value == "Destinations":
                head_cell = str(cell)

                for number in range(aws_dest_row_count - 1):
                    if "." in head_cell:
                        check_str_cell = len(head_cell[head_cell.index(".") + 1:])
                        if check_str_cell < 4:
        
                            function_four(head_cell, aws_dest_dict, count)
                            count += 1
                            
                        else:
                            function_four(head_cell, aws_dest_dict, count)
                            count += 1
                            
        count = 1

#this gets aws_code, same as above:
aws_code_row_count = 0
aws_code_count = 0
for row in aws_code:
    aws_code_row_count += 1

for row in aws_code:
    for cell in row:
        for key in aws_code_dict:
            if cell.value != None and cell.value == "Code":
                head_cell = str(cell)

                for number in range(aws_code_row_count - 1):
                    if "." in head_cell:
                        check_str_cell = len(head_cell[head_cell.index(".") + 1:])
                        if check_str_cell < 4:
        
                            function_four(head_cell, aws_code_dict, count)
                            count += 1
                            
                        else:
                            function_four(head_cell, aws_code_dict, count)
                            count += 1
                            
        count = 1

#this gets ripi_destinations, same as aws_destinations:
ripi_dest_row_count = 0
ripi_dest_count = 0
for row in ripi_destinations:
    ripi_dest_row_count += 1

for row in ripi_destinations:
    for cell in row:
        for key in ripi_dest_dict:
            if cell.value != None and cell.value == "Below are all RIPI":
                head_cell = str(cell)

                for number in range(ripi_dest_row_count - 1):
                    if "." in head_cell:
                        check_str_cell = len(head_cell[head_cell.index(".") + 1:])
                    
                        if check_str_cell < 16:
        
                            function_five(head_cell, ripi_dest_dict, count)
                            count += 1
                            
                        else:
                            function_four(head_cell, ripi_dest_dict, count)
                            count += 1
                            
        count = 1

#this gets ripi_code, same as aws_code:
ripi_code_row_count = 0
ripi_code_count = 0
for row in ripi_code:
    ripi_code_row_count += 1
for row in ripi_code:
    for cell in row:
        for key in ripi_code_dict:
            if cell.value == None:
                head_cell = str(cell)

                for number in range(ripi_code_row_count - 1):
                    if "." in head_cell:
                        check_str_cell = len(head_cell[head_cell.index(".") + 1:])
                        if check_str_cell < 16:
        
                            function_five(head_cell, ripi_code_dict, count)
                            count += 1
                            
                        else:
                            function_four(head_cell, ripi_code_dict, count)
                            count += 1
                            
        count = 1

#this gets ipi_destinations, same as aws_destinations and ripi_destinations:
ipi_dest_row_count = 0
ipi_dest_count = 0
for row in ipi_destinations:
    ipi_dest_row_count += 1

for row in ipi_destinations:
    for cell in row:
        for key in ipi_dest_dict:
            if cell.value != None and cell.value == "Below are all IPI/MLB":
                head_cell = str(cell)

                for number in range(ipi_dest_row_count - 1):
                    if "." in head_cell:
                        check_str_cell = len(head_cell[head_cell.index(".") + 1:])
                        if check_str_cell < 16:
        
                            function_five(head_cell, ipi_dest_dict, count)
                            count += 1
                            
                        else:
                            function_four(head_cell, ipi_dest_dict, count)
                            count += 1
                            
        count = 1

#this gets ipi_code, same as aws_code and ripi_code:
ipi_code_row_count = 0
ipi_code_count = 0
for row in ipi_code:
    ipi_code_row_count += 1

for row in ipi_code:
    for cell in row:
        for key in ipi_code_dict:
            if cell.value == None:
                head_cell = str(cell)

                for number in range(ipi_code_row_count - 1):
                    if "." in head_cell:
                        check_str_cell = len(head_cell[head_cell.index(".") + 1:])
                        if check_str_cell < 16:
        
                            function_five(head_cell, ipi_code_dict, count)
                            count += 1
                            
                        else:
                            function_four(head_cell, ipi_code_dict, count)
                            count += 1
                            
        count = 1

aws_row_count = 0
aws_count = 5
for row in aws_cost:
    aws_row_count += 1

#this gets the AWS cost
for row in aws_cost:
    for cell in row:
        for key in aws_cost_dict:
            if cell.value != None and cell.value == key:

                head_next_cell = function_six(cell, aws_count)
                
                head_cell_value = head_next_cell.value
                str_head_next_cell = str(head_next_cell)
                
                for number in range(aws_row_count - 6):
                    if "." in str_head_next_cell:
                                check_str_cell = len(str_head_next_cell[str_head_next_cell.index(".") + 1:])
                                if check_str_cell < 4:
                
                                    function_four(str_head_next_cell, aws_cost_dict, count)
                                    count += 1
                                    
                                else:
                                    function_four(str_head_next_cell, aws_cost_dict, count)
                                    count += 1
                                    
                count = 1                                

ripi_row_count = 0
ripi_count = 31
for row in ripi_cost:
    ripi_row_count += 1

#this gets ripi_cost
for row in ripi_cost:
    for cell in row:
        for key in ripi_cost_dict:
            if cell.value != None and cell.value == key:

                head_next_cell = function_six(cell, ripi_count)
                
                head_cell_value = head_next_cell.value
                str_head_next_cell = str(head_next_cell)
                
                for number in range(ripi_row_count - 32):
                    if "." in str_head_next_cell:
                                check_str_cell = len(str_head_next_cell[str_head_next_cell.index(".") + 1:])
                                if check_str_cell < 4:
                                    function_four(str_head_next_cell, ripi_cost_dict, count)
                                    count += 1
                                else:
                                    function_five(str_head_next_cell, ripi_cost_dict, count)
                                    count += 1                                   
                count = 1

ipi_row_count = 0
ipi_count = 56
for row in ipi_cost:
    ipi_row_count += 1

for row in ipi_cost:
    for cell in row:
        for key in ipi_cost_dict:
            if cell.value != None and cell.value == key:
                head_next_cell = function_six(cell, ipi_count)
                
                head_cell_value = head_next_cell.value
                str_head_next_cell = str(head_next_cell)
                
                for number in range(ipi_row_count - 57):
                    if "." in str_head_next_cell:
                                check_str_cell = len(str_head_next_cell[str_head_next_cell.index(".") + 1:])
                                if check_str_cell < 4:
                
                                    function_four(str_head_next_cell, ipi_cost_dict, count)
                                    count += 1
                                    
                                else:
                                    function_five(str_head_next_cell, ipi_cost_dict, count)
                                    count += 1
                                    
                count = 1
                
print("HEADER: " + str(header_dict) + "\n")
print("SELL PRICE: " + str(sell_dict) + "\n")
print("COST PRICE: " + str(of_dict) + "\n")
print("DRAY: " + str(dray_dict) + "\n")
print("ARB: " + str(arb_dict) + "\n")
print("Sheet Header: " + str(header_dict) + "\n")
print("Renewal Date: " + str(renew_date_dict) + "\n")
print("Reps:\n" + str(rep_dict) + "\n")
print("Carriers:\n" + str(carrier_dict) + "\n")
print("Effective Date: " + str(effect_date_dict) + "\n")
print("Expiration Date: " + str(expire_date_dict) + "\n")
print("Comm Bullet: " + str(comm_bullet_dict) + "\n")
print("AWS Cost: " + str(aws_cost_dict) + "\n")
print("Ripi Cost: " + str(ripi_cost_dict) + "\n")
print("Ipi Cost: " + str(ipi_cost_dict) + "\n")
print("AWS Destinations: " + str(aws_dest_dict) + "\n")
print("AWS Codes: " + str(aws_code_dict) + "\n")
print("Ripi Destinations: " + str(ripi_dest_dict) + "\n")
print("Ripi Codes: " + str(ripi_code_dict) + "\n")
print("Ipi Destinations: " + str(ipi_dest_dict) + "\n")
print("Ipi Codes: " + str(ipi_code_dict) + "\n")


#open a new notepad file and save to "C:\Users\amonroy.lax\Documents\dev\py_text"
with open("C:/Users/amonroy.lax/Documents/dev/py_text/readmeCOPY.txt", "w") as f:
    f.write("HEADER: " + str(header_dict) + "\n")
    f.write("\n")
    f.write("SELL PRICE: " + str(sell_dict) + "\n")
    f.write("\n")
    f.write("COST PRICE: " + str(of_dict) + "\n")
    f.write("\n")
    f.write("DRAY: " + str(dray_dict) + "\n")
    f.write("\n")
    f.write("ARB: " + str(arb_dict) + "\n")
    f.write("\n")
    f.write("Sheet Header: " + str(header_dict) + "\n")
    f.write("\n")
    f.write("Renewal Date: " + str(renew_date_dict) + "\n")
    f.write("\n")
    f.write("Reps:\n" + str(rep_dict) + "\n")
    f.write("\n")
    f.write("Carriers:\n" + str(carrier_dict) + "\n")
    f.write("\n")
    f.write("Effective Date: " + str(effect_date_dict) + "\n")
    f.write("\n")
    f.write("Expiration Date: " + str(expire_date_dict) + "\n")
    f.write("\n")
    f.write("Comm Bullet: " + str(comm_bullet_dict) + "\n")
    f.write("\n")
    f.write("AWS Cost: " + str(aws_cost_dict) + "\n")
    f.write("\n")
    f.write("Ripi Cost: " + str(ripi_cost_dict) + "\n")
    f.write("\n")
    f.write("Ipi Cost: " + str(ipi_cost_dict) + "\n")
    f.write("\n")
    f.write("AWS Destinations: " + str(aws_dest_dict) + "\n")
    f.write("\n")
    f.write("AWS Codes: " + str(aws_code_dict) + "\n")
    f.write("\n")
    f.write("Ripi Destinations: " + str(ripi_dest_dict) + "\n")
    f.write("\n")
    f.write("Ripi Codes: " + str(ripi_code_dict) + "\n")
    f.write("\n")
    f.write("Ipi Destinations: " + str(ipi_dest_dict) + "\n")
    f.write("\n")
    f.write("Ipi Codes: " + str(ipi_code_dict) + "\n")
    f.write("\n")
    f.write("END")
    f.write("\n")

    f.close()

#save the above notepad file to "C:\Users\amonroy.lax\Documents\dev\py_text"
with open("C:/Users/amonroy.lax/Documents/dev/py_text/readmeCOPY.txt", "r") as f:
    data = f.read()
    lines = data.splitlines()
    words = data.split()
    len_lines = len(lines)
    len_words = len(words)
    str_lines = str(len_lines)
    str_words = str(len_words)
    f.close()

    print("Number of lines: " + str_lines)
    print("Number of words: " + str_words)

with open("C:/Users/amonroy.lax/Documents/dev/py_text/readmeCOPY.txt", "a") as f:
    f.write("Number of lines: " + str_lines)
    f.write("\n")
    f.write("Number of words: " + str_words)
    f.write("\n")
    f.close()

print("END")



