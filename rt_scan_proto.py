from openpyxl import load_workbook
import time
import datetime

print("STARTING")

#open and load file, select OF sheet, search header portion
wb = load_workbook(filename='Acme Furniture 2022 Feb 1- 14.xlsx', data_only=True)
sbs = load_workbook(filename= 'sbs.xlsx', data_only=True)
ws = wb.active
sbs_ws = sbs["1. CBP_40HC"]
sheet_headers = ws["A4:J9"]
fcl_table = ws["A10:J27"]
of_table = ws["L10:AB26"]
dray_table = ws["N10:R26"]
arb_table = ws["X10:AC27"]


#list of header dictionaries

header_list = ["Attn:", "Company:", "Company Email:", "From:", "Date:", "Email:", "EFFECTIVE GATE IN:", "to", "COMMODITY:"]

header_dict = {"Attn": None, "Company": None, "Company_Email": None, "From": None, "Date": None, "Email": None, "Effective_Gate_In": None, "Expires": None, "COMMODITY": None}
dray_dict = {"Chassis": [], "I/H": [], "I/H Inclusive notes": [], "Live unload / Drop Pick": [], "Drayage quote# / expire date": []}
of_dict = {"CARRIERS": [], "CY or RAMP": [], "COST 20'": [], "COST 40'": [], "COST 40HQ": [], "COST 45'": [], "PROFIT": [], "Other Per Cntr": []}
arb_dict = {"ARB COST 20'": [], "ARB COST 40'": [], "ARB COST 40HQ": [], "ARB COST 45'": [], "REMARKS": []}
sell_dict = {"ORIGIN": [], "DESTINATION": [], "MODE": [], "TIER": [], "CHASSIS": [], "20'": [], "40'": [], "HQ'": [], "45'": [], "REMARKS": []}


for row in sheet_headers:
    for cell in row:
        if cell.value != None and cell.value in header_list:
            #print(cell.value)
            str_cell = str(cell)

                        #Convert col letter to int equivalent
            col_pos = ord(str_cell[-3])

                        #get row position
            row_pos = str_cell[-2]

                        #get next column position
            next_col_pos = chr(col_pos + 1)

                        #get next cell position
            next_cell = ws[(next_col_pos + row_pos)]

                        #returns <CELL '020123'.B5> as string. I think i was going to use this to parse. Not sure anymore.
                        #str_next_cell = str(next_cell)

                        #Store the value of next_cell, which should be value of the header
            cell_value = next_cell.value

            #check whether cell.value matches dictionary key
            for key in header_dict:
                cell_append = cell.value[:-1]
                #case switch statement that checks if cell.value matches "Company_Email" or "Effective_Gate_In" or "Expires"
                if cell.value == "Company Email:":
                    header_dict.update({"Company_Email": next_cell.value})
                elif cell.value == "EFFECTIVE GATE IN:":
                    cell_date_time = str(next_cell.value)
                    header_dict.update({"Effective_Gate_In": cell_date_time})
                elif cell.value == "to":
                    cell_date_time = str(next_cell.value)
                    header_dict.update({"Expires": cell_date_time})
                elif cell_append == key:
                    #print(key)
                    header_dict.update({key: next_cell.value})
                elif cell.value == "Date:":
                    cell_date_time = str(next_cell.value)
                    header_dict.update({"Date": cell_date_time})
            #print(cell_value)
#row stuff
count = 1 
row_amount = 0
row_count = 0
dray_row_amount = 0
arb_row_amount = 0        
for row in fcl_table:
    row_amount += 1
for row in dray_table:
    dray_row_amount += 1
for row in arb_table:
    arb_row_amount += 1

#sell_rate section
for row in fcl_table:
    for cell in row:
        for key in sell_dict:
            for number in range(row_amount - 2):
                if cell.value != None and cell.value == key:
                    #checks for chassis header
                    if key == "CHASSIS":
                        #print("found it")
                        str_cell = str(cell)
                        #print(str_cell)
                        col_pos = str_cell[-4]
                        #print(col_pos)
                        row_pos = str_cell[-3:-1]
                        #print(row_pos)
                        next_row_pos = str(int(row_pos) + count)
                        #print(next_row_pos)
                        next_cell = ws[(col_pos + next_row_pos)]
                        #print(next_cell)
                        cell_value = next_cell.value
                        #print(cell_value)

                        #section below gets the number of chassis days
                        chassis_str_cell = str(cell)
                        #print(chassis_str_cell)
                                            #Convert col letter to int equivalent
                        chassis_col_pos = ord(chassis_str_cell[-4])
                        #print(chassis_col_pos)
                                            #get row position
                        chassis_row_pos = chassis_str_cell[-3:-1]
                        #print(chassis_row_pos)

                        chassis_next_row_pos = str(int(chassis_row_pos) + count)
                        #print(chassis_next_row_pos)
                                            #get next column position
                        chassis_next_col_pos = chr(chassis_col_pos + 1)
                        #print(chassis_next_col_pos)
                                            #get next cell position
                        chassis_next_cell = ws[(chassis_next_col_pos + chassis_next_row_pos)]
                        #print(chassis_next_cell)

                        chassis_cell_value = chassis_next_cell.value[:6]
                        #print(chassis_cell_value)
                        chassis = cell_value +' '+ chassis_cell_value

                        sell_dict[key].append(chassis)
                        count += 1
                    
                    else:
                        #sell_dict[key].append(cell_value)
                        str_cell = str(cell)
                        #print(str_cell)
                        col_pos = str_cell[-4]
                        #print(col_pos)
                        row_pos = str_cell[-3:-1]
                        #print(row_pos)
                        next_row_pos = str(int(row_pos) + count)
                        #print(next_row_pos)
                        next_cell = ws[(col_pos + next_row_pos)]
                        #print(next_cell)
                        cell_value = next_cell.value
                        #print(cell_value)
                        sell_dict[key].append(cell_value)
                        count += 1
                        #print("done")
                
            count = 1

#this section is for cost
for row in of_table:
    for cell in row:
        for key in of_dict:
            for number in range(row_amount - 2):
                if cell.value != None and cell.value == key:
                    #print("done")
                    str_cell = str(cell)
                    #search for "." in str_cell
                    if "." in str_cell:
                        check_str_cell = len(str_cell[str_cell.index(".") + 1:])
                        if check_str_cell > 4:

                            #print(str_cell)
                            col_pos = str_cell[-5:-3]
                            #print(col_pos)
                            row_pos = str_cell[-3:-1]
                            #print(row_pos)
                            next_row_pos = str(int(row_pos) + count)
                            #print(next_row_pos)
                            next_cell = ws[(col_pos + next_row_pos)]
                            #print(next_cell)
                            cell_value = next_cell.value
                            #print(cell_value)
                            of_dict[key].append(cell_value)
                            count += 1
                        else:
                            #print(str_cell)
                            col_pos = str_cell[-4]
                            #print(col_pos)
                            row_pos = str_cell[-3:-1]
                            #print(row_pos)
                            next_row_pos = str(int(row_pos) + count)
                            #print(next_row_pos)
                            next_cell = ws[(col_pos + next_row_pos)]
                            #print(next_cell)
                            cell_value = next_cell.value
                            #print(cell_value)
                            of_dict[key].append(cell_value)
                            count += 1
            count = 1

#this section is for dray
for row in dray_table:
    for cell in row:
        for key in dray_dict:
            for number in range(dray_row_amount - 2):
                if cell.value != None and cell.value == key:
                    str_cell = str(cell)
                    #print(str_cell)
                    col_pos = str_cell[-4]
                    #print(col_pos)
                    row_pos = str_cell[-3:-1]
                    #print(row_pos)
                    next_row_pos = str(int(row_pos) + count)
                    #print(next_row_pos)
                    next_cell = ws[(col_pos + next_row_pos)]
                    #print(next_cell)
                    cell_value = next_cell.value
                    #print(cell_value)
                    dray_dict[key].append(cell_value)
                    count += 1
            count = 1

#this section is for arb
for row in arb_table:
    for cell in row:
        for key in arb_dict:
            for number in range(arb_row_amount - 2):
                if cell.value != None and cell.value == key:
                    #print("done")
                    str_cell = str(cell)
                    #search for "." in str_cell
                    if "." in str_cell:
                        check_str_cell = len(str_cell[str_cell.index(".") + 1:])
                        if check_str_cell > 4:

                            #print(str_cell)
                            col_pos = str_cell[-5:-3]
                            #print(col_pos)
                            row_pos = str_cell[-3:-1]
                            #print(row_pos)
                            next_row_pos = str(int(row_pos) + count)
                            #print(next_row_pos)
                            next_cell = ws[(col_pos + next_row_pos)]
                            #print(next_cell)
                            cell_value = next_cell.value
                            #print(cell_value)
                            arb_dict[key].append(cell_value)
                            count += 1
                        else:
                            #print(str_cell)
                            col_pos = str_cell[-4]
                            #print(col_pos)
                            row_pos = str_cell[-3:-1]
                            #print(row_pos)
                            next_row_pos = str(int(row_pos) + count)
                            #print(next_row_pos)
                            next_cell = ws[(col_pos + next_row_pos)]
                            #print(next_cell)
                            cell_value = next_cell.value
                            #print(cell_value)
                            arb_dict[key].append(cell_value)
                            count += 1
            count = 1

print("HEADER")
print(header_dict)
print("SELL")
print(sell_dict)
print("COST")
print(of_dict)
print("DRAY")
print(dray_dict)
print("ARB")
print(arb_dict)
print("EOP")

#add an additional value to header_dict["Carriers"] list
#header_dict["Carriers"].append("Carrier4")
#print(header_dict)